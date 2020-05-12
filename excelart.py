import openpyxl
from PIL import Image
from PIL import GifImagePlugin
import glob
import os


# Image to colored Excel cells
def image_to_excel(file, output, percentage, zoom):

    # Open picture and create a workbook instance
    im = Image.open(file)
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Resize image with spreadsheet's columns
    width, height = im.size
    cols = width * percentage/100
    print('Old image size: ' + str(im.size))
    imgScale = cols/width
    newSize = (int(width*imgScale), int(height*imgScale))
    im = im.resize(newSize)

    # Get new picture's dimensions
    cols, rows = im.size
    print('New image size: ' + str(im.size))

    # Formatting spreadsheet's cell size: height = 0.6 and width = 0.0625 (1 pixel)
    zoom = 100
    for i in range(1, rows):
        sheet.row_dimensions[i].height = zoom*0.6/100
    for j in range(1, cols):
        column_letter = openpyxl.utils.get_column_letter(j)
        sheet.column_dimensions[column_letter].width = zoom*0.0625/100

    # Convert image to RGB
    rgb_im = im.convert('RGB')

    # Formatting cell's color 
    for i in range(1, rows):
        for j in range(1, cols):
            c = rgb_im.getpixel((j, i))
            rgb2hex = lambda r,g,b: f"{r:02x}{g:02x}{b:02x}"
            c = rgb2hex(*c)
            sheet.cell(row = i, column = j).value = " "
            customFill = openpyxl.styles.PatternFill(start_color=c, end_color=c, fill_type='solid')
            sheet.cell(row = i, column = j).fill = customFill

    # Save workbook
    wb.save(output) 
    wb.close()
    
# Convert a gif image to its frame images - saved in output_folder
def gif_to_img(file, output_folder, step):
    
    # Creating folder
    try: 
        os.mkdir(output_folder) 
    except OSError as error: 
        print(error)  
    
    # Open Gif Image
    imageObject = Image.open(file)

    # Number of frames
    print('Total number of frames: ' + str(round(imageObject.n_frames/step)))

    # Save all frames as a image
    for frame in range(0, imageObject.n_frames, step):
        imageObject.seek(frame)
        imageObject.convert('RGB').save(output_folder+'//'+os.path.splitext(file)[0]+str(frame)+'.jpg')
        
# Save each frame on a excel's sheet
def gif_to_excel(images, output, percentage, zoom):
    wb = openpyxl.Workbook()
    
    for frame in range(len(images)):
        # Open picture and create a workbook instance
        im = Image.open(images[frame])
        sheet = wb.create_sheet('nazare' + str(frame))

        # Resize image with spreadsheet's columns
        width, height = im.size
        cols = width * percentage/100
        print('Old image size: ' + str(im.size))
        imgScale = cols/width
        newSize = (int(width*imgScale), int(height*imgScale))
        im = im.resize(newSize)

        # Get new picture's dimensions
        cols, rows = im.size
        print('New image size: ' + str(im.size))
        print('Frame: ' + str(frame))

        # Formatting spreadsheet's cell size: height = 0.6 and width = 0.0625 (1 pixel)
        zoom = 100
        for i in range(1, rows):
            sheet.row_dimensions[i].height = zoom*0.6/100
        for j in range(1, cols):
            column_letter = openpyxl.utils.get_column_letter(j)
            sheet.column_dimensions[column_letter].width = zoom*0.0625/100

        # Convert image to RGB
        rgb_im = im.convert('RGB')

        # Formatting cell's color 
        for i in range(1, rows):
            for j in range(1, cols):
                c = rgb_im.getpixel((j, i))
                rgb2hex = lambda r,g,b: f"{r:02x}{g:02x}{b:02x}"
                c = rgb2hex(*c)
                sheet.cell(row = i, column = j).value = " "
                customFill = openpyxl.styles.PatternFill(start_color=c, end_color=c, fill_type='solid')
                sheet.cell(row = i, column = j).fill = customFill

    # Save workbook
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(output) 
    wb.close()
    print('Finalizado !')
    
    

# Export (input file, output file, image scale (%), zoom(%))
image_to_excel('jangada.jpg', 'jangada.xlsx', 100, 100)

# gif_to_img(file.gif, output_folder_name, step_frames) 
print('\n')
gif_to_img('nazare.gif', 'nazare_photos', 2)

# Folder "nazare_photos" with all frames in *.jpg
print('\n')
images = glob.glob("nazare_photos/*.jpg")
gif_to_excel(images, 'nazare.xlsx', 100, 100)
